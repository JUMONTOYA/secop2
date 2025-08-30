import requests
import pandas as pd
from io import StringIO, BytesIO
from urllib.parse import quote
from datetime import datetime
from pathlib import Path
import logging
import time
from flask import Flask, request, jsonify, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# --- Configuración de la aplicación Flask ---
# El argumento static_folder='.' le dice a Flask que sirva archivos desde el directorio actual
app = Flask(__name__, static_folder='.')

# --- Lógica del Script Original (adaptada a funciones) ---

# Configuración del logging
log_dir = Path.home() / "secop2_logs_web"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / f"secop2_web_log_{datetime.now().strftime('%Y%m%d')}.log"
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s',
                    handlers=[logging.FileHandler(log_file, encoding='utf-8')])

def escapar_sql_mejorado(texto):
    if not texto: return texto
    return texto.replace("'", "''").strip()

def procesar_terminos_multiples(entrada):
    if not entrada: return []
    return [escapar_sql_mejorado(termino.strip()) for termino in entrada.split(',') if termino.strip()]

def construir_consulta_where(filtros):
    condiciones = []
    
    if filtros.get('proceso_de_compra'):
        numeros_proceso = procesar_terminos_multiples(filtros['proceso_de_compra'])
        if numeros_proceso:
            lista_formateada = ", ".join([f"'{n}'" for n in numeros_proceso])
            condiciones.append(f"proceso_de_compra IN ({lista_formateada})")
            
    campos_multiples = {
        'entidad': 'entidad',
        'departamento': 'departamento_entidad',
        'ciudad': 'ciudad_entidad',
        'modalidades': 'modalidad_de_contratacion'
    }
    for clave, campo_sql in campos_multiples.items():
        if filtros.get(clave):
            terminos = procesar_terminos_multiples(filtros[clave])
            if terminos:
                condiciones_or = [f"upper({campo_sql}) like upper('%{termino}%')" for termino in terminos]
                condiciones.append(f"({' OR '.join(condiciones_or)})")
    
    if filtros.get('estado_del_procedimiento'):
        estado = escapar_sql_mejorado(filtros['estado_del_procedimiento'])
        condiciones.append(f"estado_del_procedimiento = '{estado}'")

    if filtros.get('estado_de_apertura_del_proceso'):
        estado_apertura = escapar_sql_mejorado(filtros['estado_de_apertura_del_proceso'])
        condiciones.append(f"estado_de_apertura_del_proceso = '{estado_apertura}'")

    if filtros.get('fecha_inicio'):
        condiciones.append(f"fecha_de_publicacion_del >= '{filtros['fecha_inicio']}T00:00:00'")
    if filtros.get('fecha_fin'):
        condiciones.append(f"fecha_de_publicacion_del <= '{filtros['fecha_fin']}T23:59:59'")
    
    return " AND ".join(condiciones)

def descargar_datos_paginado(where_clause, max_registros=50000):
    logging.info(f"Iniciando descarga paginada. Límite: {max_registros}.")
    todas_las_paginas, offset, total_registros = [], 0, 0
    limite_por_pagina = 10000 

    while total_registros < max_registros:
        BASE_URL = "https://www.datos.gov.co/resource/p6dx-8zbt.csv"
        limite_actual = min(limite_por_pagina, max_registros - total_registros)
        query = f"$limit={limite_actual}&$offset={offset}"
        if where_clause:
            query += f"&$where={quote(where_clause)}"
        
        url = f"{BASE_URL}?{query}"
        logging.debug(f"Consultando URL: {url}")
        
        response = requests.get(url, timeout=90)
        response.raise_for_status()
        
        df_pagina = pd.read_csv(StringIO(response.text))
        if df_pagina.empty:
            break
            
        todas_las_paginas.append(df_pagina)
        registros_pagina = len(df_pagina)
        total_registros += registros_pagina
        
        if registros_pagina < limite_por_pagina:
            break
        offset += limite_por_pagina
        
    if not todas_las_paginas:
        return pd.DataFrame()

    return pd.concat(todas_las_paginas, ignore_index=True)


def crear_excel_en_memoria(df):
    output = BytesIO()
    wb = Workbook()
    
    ws_datos = wb.active
    ws_datos.title = "Datos SECOP2"
    
    font_titulo_empresa = Font(bold=True, size=16, color="003366")
    fill_titulo_empresa = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    alignment_titulo_empresa = Alignment(horizontal='center', vertical='center')
    
    ws_datos['A1'] = "Sociedad Colombiana de Consultoria SAS"
    ws_datos['A1'].font = font_titulo_empresa
    ws_datos['A1'].fill = fill_titulo_empresa
    ws_datos['A1'].alignment = alignment_titulo_empresa
    ws_datos.merge_cells('A1:J1')
    ws_datos.row_dimensions[1].height = 30

    column_names = list(df.columns)
    ws_datos.append(column_names)

    for row in dataframe_to_rows(df, index=False, header=False):
        ws_datos.append(row)
    
    tab = Table(displayName="TablaDatos", ref=f"A2:{get_column_letter(ws_datos.max_column)}{ws_datos.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws_datos.add_table(tab)

    for column_cells in ws_datos.columns:
        if isinstance(column_cells[0], tuple): 
            continue
        column_letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws_datos.column_dimensions[column_letter].width = length + 2
    
    wb.save(output)
    output.seek(0)
    return output

# --- Rutas de la API ---

@app.route('/')
def index():
    # Sirve el archivo con el nuevo nombre
    return send_from_directory(app.static_folder, 'herramienta-secop.html')

@app.route('/api/query', methods=['POST'])
def handle_query():
    fecha_limite = datetime(2025, 12, 31)
    if datetime.now() > fecha_limite:
        logging.warning("Intento de acceso después de la fecha de expiración.")
        return jsonify({"error": "Este servicio ha expirado."}), 403

    try:
        filtros = request.json
        logging.info(f"Recibida nueva consulta con filtros: {filtros}")

        where_clause = construir_consulta_where(filtros)
        max_registros = int(filtros.get('max_registros', 50000))
        
        df = descargar_datos_paginado(where_clause, max_registros=max_registros)

        if df.empty:
            logging.warning("No se encontraron registros.")
            return jsonify({"error": "No se encontraron registros con los filtros especificados."}), 404
        
        excel_file = crear_excel_en_memoria(df)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"secop2_reporte_{timestamp}.xlsx"

        logging.info(f"Enviando archivo: {filename} con {len(df)} registros.")
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except requests.exceptions.HTTPError as e:
        logging.error(f"Error HTTP de la API de Datos Abiertos: {e}")
        return jsonify({"error": f"Error al contactar la API de SECOP: {e.response.text}"}), 502
    except Exception as e:
        logging.error(f"Error inesperado en el servidor: {e}", exc_info=True)
        return jsonify({"error": "Ocurrió un error inesperado en el servidor."}), 500

# Esta parte solo se usa para pruebas locales.
# En producción, el servidor Gunicorn inicia la app.
if __name__ == '__main__':
    app.run(debug=True, port=5000)
